"""
xlsm_to_js.py
Reads PPS Excel workbook and overwrites pps-data.js.
Preserves PPS_2, PROJECTS, and USERS blocks already in the file.
"""

import re
import sys
import json
from datetime import datetime, timedelta
from pathlib import Path
import openpyxl

# ── Config ─────────────────────────────────────────────────────────────────────
XLSM_PATH = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("PPS_1450_Frazee_Sep_24_2025_Reference.xlsm")
JS_PATH   = Path(sys.argv[2]) if len(sys.argv) > 2 else Path("pps-data.js")

EXCEL_EPOCH = datetime(1899, 12, 30)

# ── Helpers ────────────────────────────────────────────────────────────────────

def xl_date(val):
    """Convert Excel serial date to 'Mon DD, YYYY' string, or return '' if blank."""
    if not val:
        return ""
    try:
        n = int(float(val))
        if n < 1:
            return ""
        d = EXCEL_EPOCH + timedelta(days=n)
        return d.strftime("%b %d, %Y")
    except (ValueError, TypeError):
        return str(val) if val else ""

def clean(val):
    """Normalise a cell value to a plain Python scalar."""
    if val is None:
        return ""
    s = str(val).strip()
    return "" if s in ("None", "N/A", "nan") else s

def rows_for_sheet(wb, sheet_name):
    """Return list-of-lists of cell values for a sheet, skipping leading blank rows."""
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    # Drop rows that are entirely empty or start with None x5 (leading blank cols)
    data = []
    for row in all_rows:
        stripped = [c for c in row if c is not None and str(c).strip()]
        if stripped:
            data.append(list(row))
    return data

def find_header_row(rows, marker):
    """Return (header_list, data_rows) by locating the row containing `marker`."""
    for i, row in enumerate(rows):
        cells = [clean(c) for c in row]
        if marker in cells:
            idx = cells.index(marker)
            headers = [clean(c) for c in row[idx:]]
            data_rows = []
            for r in rows[i+1:]:
                vals = [clean(c) for c in r[idx:]]
                if any(v for v in vals):
                    data_rows.append(vals)
            return headers, data_rows
    return [], []

def row_to_dict(headers, row):
    d = {}
    for i, h in enumerate(headers):
        d[h] = row[i] if i < len(row) else ""
    return d

def js_str(v):
    """Escape a value as a JS string literal."""
    return json.dumps(str(v) if v is not None else "")

def js_num(v, default=0):
    try:
        f = float(v)
        return int(f) if f == int(f) else round(f, 2)
    except (ValueError, TypeError):
        return default

def js_bool(v):
    return "true" if str(v).strip().lower() in ("yes", "true", "1") else "false"

# ── Sheet parsers ──────────────────────────────────────────────────────────────

def parse_project(wb):
    """Pull project metadata from the Settings sheet."""
    ws = wb["Settings"]
    rows = list(ws.iter_rows(values_only=True))
    meta = {}
    for row in rows:
        cells = [c for c in row if c is not None]
        if len(cells) >= 2:
            key = str(cells[0]).strip()
            val = str(cells[1]).strip()
            if key == "Project Name":
                meta["name"] = val
    # Fallback: pull from first cell that contains the project name string
    if not meta.get("name"):
        for row in rows:
            for cell in row:
                if cell and "1450 Frazee" in str(cell):
                    meta["name"] = "SDCCD 1450 Frazee"
                    break

    # Pull financials summary from Dashboard for budget/plannedSpend
    try:
        dws = wb["Dashboard"]
        drows = list(dws.iter_rows(values_only=True))
        for row in drows:
            cells = [c for c in row if c is not None]
            for i, c in enumerate(cells):
                if str(c).strip() == "FUNDS REMAINING" and i + 1 < len(cells):
                    try:
                        meta["budget"] = int(float(cells[i+1]))
                    except (ValueError, TypeError):
                        pass
                if str(c).strip() == "Total Spend" and i + 1 < len(cells):
                    try:
                        meta["plannedSpend"] = int(float(cells[i+1]))
                    except (ValueError, TypeError):
                        pass
    except Exception:
        pass

    return meta

def parse_ffe(wb):
    rows = rows_for_sheet(wb, "Master FF&E List")
    headers, data = find_header_row(rows, "FF&E ID")
    items = []
    for row in data:
        d = row_to_dict(headers, row)
        fid = clean(d.get("FF&E ID", ""))
        if not fid.startswith("FFE-"):
            continue
        cur_bld = clean(d.get("Current Building Location", ""))
        cur_fl  = clean(d.get("Current Floor Location", ""))
        new_bld = clean(d.get("New Building Location", ""))
        new_fl  = clean(d.get("New Floor Location", ""))
        loc_cur = f"{cur_bld} · {cur_fl}" if cur_bld and cur_fl else cur_bld or cur_fl
        loc_new = f"{new_bld} · {new_fl}" if new_bld and new_fl else new_bld or new_fl
        items.append({
            "id":            fid,
            "name":          clean(d.get("Item Name", "")),
            "desc":          clean(d.get("Item Description", "")),
            "type":          clean(d.get("Product Type", "")),
            "vendor":        clean(d.get("Vendor/Supplier", "")),
            "vendorPOC":     clean(d.get("Vendor/Supplier POC", "")),
            "userGroup":     clean(d.get("User/User Group", "")),
            "dimensions":    clean(d.get("Ext. Dimensions", "")),
            "power":         clean(d.get("Power Type Required", "")),
            "data":          clean(d.get("Data Type Required", "")),
            "mounting":      clean(d.get("Mounting Type", "")),
            "locationCurrent": loc_cur,
            "locationNew":   loc_new,
            "roomTypeCurrent": clean(d.get("Current Room Type", "")),
            "roomNoCurrent": clean(d.get("Current Room No.", "")),
            "roomType":      clean(d.get("New Room Type", "")),
            "roomNo":        clean(d.get("New Room No.", "")),
            "procMethod":    clean(d.get("Procurement Method & Installation", "")),
            "fundingSource": clean(d.get("Funding Source", "")),
            "qtyNeeded":     js_num(d.get("Total Item Quantity Needed", 0)),
            "qtyExisting":   js_num(d.get("Existing Quantity (If Re-used)", 0)),
            "qtyNew":        js_num(d.get("New Item Quantity (To Order)", 0)),
            "unitPrice":     js_num(d.get("Unit Price of New Item", 0)),
            "extCost":       js_num(d.get("Extended Cost of Item", 0)),
            "eta":           xl_date(d.get("ETA", "")),
            "processStatus": clean(d.get("Process Status", "")),
            "ffeStatus":     clean(d.get("FF&E Item Status", "")),
            "condition":     clean(d.get("Item Condition", "")),
            "warranty":      clean(d.get("Warranty (Y/N)", "")),
            "warrantyType":  clean(d.get("Warranty Type", "")),
            "priority":      clean(d.get("Priority", "")),
            "punchList":     clean(d.get("Punch List Item (Y/N)", "")),
            "swing":         d.get("Swing Space Item (Y/N)", ""),
            "moveId":        clean(d.get("Associated Move Management ID", "")),
        })
    return items

def parse_move(wb):
    rows = rows_for_sheet(wb, "Move Management")
    headers, data = find_header_row(rows, "Move Management ID")
    records = []
    for row in data:
        d = row_to_dict(headers, row)
        mid = clean(d.get("Move Management ID", ""))
        if not mid.startswith("MM-"):
            continue
        cur_bld = clean(d.get("Current Building Location", ""))
        cur_fl  = clean(d.get("Current Floor Location", ""))
        new_bld = clean(d.get("New Building Location", ""))
        new_fl  = clean(d.get("New Floor Location", ""))
        records.append({
            "id":        mid,
            "itemName":  clean(d.get("Items Being Moved", "")),
            "userGroup": clean(d.get("User/User Group", "")),
            "fromBldg":  cur_bld,
            "fromFloor": cur_fl,
            "fromRoom":  clean(d.get("Current Room No.", "")),
            "toBldg":    new_bld,
            "toFloor":   new_fl,
            "toRoom":    clean(d.get("New Room No.", "")),
            "roomType":  clean(d.get("New Room Type", "")),
            "moveDate":  xl_date(d.get("Scheduled Move Date", "")),
            "dateMoved": xl_date(d.get("Date Moved to New Location", "")),
            "mover":     clean(d.get("Transportation Entity", "")),
            "moverPOC":  clean(d.get("Transportation Entity POC", "")),
            "status":    clean(d.get("Move Management Status", "")),
            "moveCost":  js_num(d.get("Move Cost", 0)),
            "swing":     d.get("Swing Space Item (Y/N)", ""),
            "ffeId":     clean(d.get("Associated FF&E ID", "")),
            "ssId":      clean(d.get("Associated Swing Space ID", "")),
            "poc":       clean(d.get("Client/Owner POC", "")),
            "notes":     clean(d.get("Notes", "")),
        })
    return records

def parse_schedule(wb):
    rows = rows_for_sheet(wb, "Schedule")
    headers, data = find_header_row(rows, "Activity")
    activities = []
    for i, row in enumerate(data):
        d = row_to_dict(headers, row)
        name = clean(d.get("Activity", ""))
        if not name or name == "0":
            continue
        activities.append({
            "id":              f"SA-{i+1}",
            "name":            name,
            "type":            clean(d.get("Activity Type", "")),
            "plannedStart":    xl_date(d.get("Planned Start", "")),
            "plannedFinish":   xl_date(d.get("Planned Finish", "")),
            "plannedDuration": js_num(d.get("Planned Duration (in Days)", 0)),
            "actualStart":     xl_date(d.get("Actual Start", "")),
            "actualFinish":    xl_date(d.get("Actual Finish", "")),
            "actualDuration":  js_num(d.get("Actual Duration (in Days)", 0)),
            "variance":        js_num(d.get("Variance", 0)),
            "pctComplete":     js_num(d.get("% Complete", 0)),
        })
    return activities

def parse_swing(wb):
    rows = rows_for_sheet(wb, "Swing Space")
    headers, data = find_header_row(rows, "Swing Space ID")
    items = []
    for row in data:
        d = row_to_dict(headers, row)
        sid = clean(d.get("Swing Space ID", ""))
        if not sid.startswith("SS-"):
            continue
        items.append({
            "id":           sid,
            "userGroup":    clean(d.get("User/User Group", "")),
            "phase":        clean(d.get("Swing Space Phase", "")),
            "occupants":    clean(d.get("No. of Occupants", "")),
            "ffeNeeds":     clean(d.get("FF&E Needs", "")),
            "techNeeds":    clean(d.get("Technology Needs", "")),
            "facilityNeeds":clean(d.get("Facility/Space Needs", "")),
            "fromBldg":     clean(d.get("Current Building Location", "")),
            "fromFloor":    clean(d.get("Current Floor Location", "")),
            "fromRoomType": clean(d.get("Current Room Type", "")),
            "fromRoom":     clean(d.get("Current Room No.", "")),
            "moveToSwingDate": xl_date(d.get("Scheduled Move Date to Swing Space", "")),
            "swingBldg":    clean(d.get("Swing Building", "")),
            "swingFloor":   clean(d.get("Swing Floor", "")),
            "swingRoomType":clean(d.get("Swing Room Type", "")),
            "swingRoom":    clean(d.get("Swing Room No.", "")),
            "moveToNewDate":xl_date(d.get("Scheduled Move Date to New Location", "")),
            "newBldg":      clean(d.get("New Building Location", "")),
            "newFloor":     clean(d.get("New Floor Location", "")),
            "newRoomType":  clean(d.get("New Room Type", "")),
            "newRoom":      clean(d.get("New Room No.", "")),
            "status":       clean(d.get("Swing Space Status", "")),
            "cost":         js_num(d.get("Swing Space Cost", 0)),
            "ffeId":        clean(d.get("Associated FF&E ID", "")),
            "moveId":       clean(d.get("Associated Move Management ID", "")),
            "poc":          clean(d.get("Client/Owner POC", "")),
            "notes":        clean(d.get("Notes", "")),
        })
    return items

def parse_issues(wb):
    rows = rows_for_sheet(wb, "Issue Log")
    headers, data = find_header_row(rows, "ID")
    issues = []
    for i, row in enumerate(data):
        d = row_to_dict(headers, row)
        lid = clean(d.get("ID", ""))
        if not lid:
            continue
        issues.append({
            "id":          f"ISS-{i+1}",
            "linkedId":    lid,
            "title":       clean(d.get("Item Name", "")),
            "type":        clean(d.get("Product Type", "")),
            "issueType":   clean(d.get("Issue Type", "")),
            "status":      clean(d.get("Issue Status", "")),
            "qtyIssue":    js_num(d.get("No. of Items with Issue", 0)),
            "qtyResolved": js_num(d.get("No. of Items Resolved", 0)),
            "description": clean(d.get("Description of Issue", "")),
            "resolution":  clean(d.get("Resolution", "")),
            "vendor":      clean(d.get("Vendor/Supplier", "")),
            "ffeId":       clean(d.get("Associated FF&E ID", "")),
            "moveId":      clean(d.get("Associated Move Management ID", "")),
            "ssId":        clean(d.get("Associated Swing Space ID", "")),
            "notes":       clean(d.get("Notes", "")),
        })
    return issues

def parse_ti(wb):
    rows = rows_for_sheet(wb, "Tenant Improvement")
    headers, data = find_header_row(rows, "Tenant Improvement ID")
    items = []
    for row in data:
        d = row_to_dict(headers, row)
        tid = clean(d.get("Tenant Improvement ID", ""))
        if not tid.startswith("TI-"):
            continue
        items.append({
            "id":          tid,
            "userGroup":   clean(d.get("User/User Group", "")),
            "scope":       clean(d.get("Facility/Space Needs", "")),
            "trade":       clean(d.get("Trade", "")),
            "building":    clean(d.get("Building", "")),
            "floor":       clean(d.get("Floor", "")),
            "roomType":    clean(d.get("Room Type", "")),
            "roomNo":      clean(d.get("Room No.", "")),
            "status":      clean(d.get("Status", "")),
            "actionOwner": clean(d.get("Action Owner", "")),
            "priority":    clean(d.get("Priority", "")),
            "vendor":      clean(d.get("Vendor", "")),
            "vendorPOC":   clean(d.get("Vendor POC", "")),
            "poc":         clean(d.get("Client/Owner POC", "")),
            "plannedCost": js_num(d.get("Quoted Cost", 0)),
            "notes":       clean(d.get("Notes", "")),
        })
    return items

def parse_financials(wb):
    rows = rows_for_sheet(wb, "Dashboard")
    headers, data = find_header_row(rows, "Cost Area")
    COLOR_MAP = {
        "FF&E (Non Ed-Tech/IT)": "#4f9cf9",
        "IT":                    "#6b7591",
        "AV/Ed-Tech":            "#6b7591",
        "Other":                 "#38d9a9",
        "Move Management":       "#6b7591",
        "Swing Space":           "#6b7591",
        "Tenant Improvement":    "#ffa94d",
        "Misc. Item Fees/Taxes": "#6b7591",
        "Non-FF&E":              "#6b7591",
    }
    items = []
    for row in data:
        d = row_to_dict(headers, row)
        label = clean(d.get("Cost Area", ""))
        if not label or label == "Total Spend":
            continue
        items.append({
            "label":   label,
            "color":   COLOR_MAP.get(label, "#6b7591"),
            "planned": js_num(d.get("Planned", 0)),
            "actual":  js_num(d.get("Actual", 0)),
        })
    return items

# ── JS serialiser ──────────────────────────────────────────────────────────────

def to_js_object(d, indent=4):
    """Serialise a dict to a compact JS object literal string."""
    pad = " " * indent
    lines = []
    for k, v in d.items():
        if isinstance(v, bool):
            lines.append(f"{pad}{k}: {'true' if v else 'false'}")
        elif isinstance(v, (int, float)):
            lines.append(f"{pad}{k}: {v}")
        elif isinstance(v, str):
            # raw bool-like strings that came from Excel
            if v in ("Yes", "No", ""):
                lines.append(f"{pad}{k}: {js_str(v)}")
            else:
                lines.append(f"{pad}{k}: {js_str(v)}")
        else:
            lines.append(f"{pad}{k}: {js_str(str(v))}")
    return "{\n" + ",\n".join(lines) + "\n" + (" " * (indent - 2)) + "}"

def render_array(name, items, indent=2):
    pad = " " * indent
    if not items:
        return f"{pad}{name}: []"
    inner = f",\n{pad}  ".join(to_js_object(item, indent + 4) for item in items)
    return f"{pad}{name}: [\n{pad}  {inner}\n{pad}]"

# ── Tail extractor ─────────────────────────────────────────────────────────────

def extract_tail(js_text):
    """Return everything from `const PPS_2` onward."""
    match = re.search(r'(// =+\s*\n// PPS DATA — SECOND PROJECT)', js_text)
    if match:
        return "\n\n" + js_text[match.start():]
    # Fallback: from `const PPS_2`
    match = re.search(r'(const PPS_2\s*=)', js_text)
    if match:
        return "\n\n" + js_text[match.start():]
    return ""

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"Reading workbook: {XLSM_PATH}")
    wb = openpyxl.load_workbook(XLSM_PATH, read_only=True, data_only=True)

    project    = parse_project(wb)
    ffe        = parse_ffe(wb)
    move       = parse_move(wb)
    schedule   = parse_schedule(wb)
    swing      = parse_swing(wb)
    issues     = parse_issues(wb)
    ti         = parse_ti(wb)
    financials = parse_financials(wb)

    print(f"  FFE items:      {len(ffe)}")
    print(f"  Move records:   {len(move)}")
    print(f"  Schedule acts:  {len(schedule)}")
    print(f"  Swing items:    {len(swing)}")
    print(f"  Issues:         {len(issues)}")
    print(f"  TI items:       {len(ti)}")
    print(f"  Financials:     {len(financials)}")

    # Read existing file to preserve tail (PPS_2, PROJECTS, USERS)
    tail = ""
    if JS_PATH.exists():
        tail = extract_tail(JS_PATH.read_text(encoding="utf-8"))

    # Build the PPS block
    lines = []
    lines.append("// =============================================================")
    lines.append("// PPS DATA — Central Data Layer")
    lines.append("// Plan Perfect Systems · v2.0")
    lines.append("//")
    lines.append("// AUTO-GENERATED by .github/scripts/xlsm_to_js.py")
    lines.append(f"// Source: {XLSM_PATH.name}")
    lines.append(f"// Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}")
    lines.append("// =============================================================")
    lines.append("")
    lines.append("const PPS = {")
    lines.append("")
    lines.append("  // ── PROJECT ────────────────────────────────────────────────")
    lines.append("  project: {")
    lines.append(f"    name:         {js_str(project.get('name', 'SDCCD 1450 Frazee'))},")
    lines.append(f"    subtitle:     \"5th Floor TI\",")
    lines.append(f"    moveDate:     \"Sep 19, 2025\",")
    lines.append(f"    budget:       {project.get('budget', 330000)},")
    lines.append(f"    plannedSpend: {project.get('plannedSpend', 309028)},")
    lines.append(f"    client:       \"Molly Gardner\",")
    lines.append(f"    clientDept:   \"Facilities Dept\",")
    lines.append(f"    mover:        \"Alexander's\",")
    lines.append(f"    moverPOC:     \"Steve\"")
    lines.append("  },")
    lines.append("")
    lines.append("  // ── LOGGED-IN USER ─────────────────────────────────────────")
    lines.append("  user: {")
    lines.append("    name:    \"John Tabler\",")
    lines.append("    initials:\"JT\",")
    lines.append("    role:    \"Project Dad\"")
    lines.append("  },")
    lines.append("")

    # ffeItems
    lines.append("  // ── FF&E ITEMS ─────────────────────────────────────────────")
    lines.append("  " + render_array("ffeItems", ffe, indent=2) + ",")
    lines.append("")

    # moveRecords
    lines.append("  // ── MOVE RECORDS ───────────────────────────────────────────")
    lines.append("  " + render_array("moveRecords", move, indent=2) + ",")
    lines.append("")

    # scheduleActivities
    lines.append("  // ── SCHEDULE ACTIVITIES ────────────────────────────────────")
    lines.append("  " + render_array("scheduleActivities", schedule, indent=2) + ",")
    lines.append("")

    # financials
    lines.append("  // ── FINANCIALS ─────────────────────────────────────────────")
    lines.append("  " + render_array("financials", financials, indent=2) + ",")
    lines.append("")

    # swingSpaceItems
    lines.append("  // ── SWING SPACE ITEMS ─────────────────────────────────────")
    lines.append("  " + render_array("swingSpaceItems", swing, indent=2) + ",")
    lines.append("")

    # tenantImprovementItems
    lines.append("  // ── TENANT IMPROVEMENT ITEMS ──────────────────────────────")
    lines.append("  " + render_array("tenantImprovementItems", ti, indent=2) + ",")
    lines.append("")

    # issues
    lines.append("  // ── ISSUES ───────────────────────────────────────────────")
    lines.append("  " + render_array("issues", issues, indent=2) + ",")
    lines.append("")

    # punchListItems — keep empty, not in Excel
    lines.append("  // ── PUNCH LIST ───────────────────────────────────────────")
    lines.append("  punchListItems: [],")
    lines.append("")

    # users — static, not in Excel
    lines.append("  // ── USERS ───────────────────────────────────────────────")
    lines.append("  users: [")
    lines.append("    { id: \"USR-1\", name: \"John Tabler\",  initials: \"JT\", email: \"jtabler@sdccd.edu\",")
    lines.append("      role: \"Work Dad\", projects: [\"SDCCD 1450 Frazee\"],")
    lines.append("      status: \"Active\", lastActive: \"Today\" },")
    lines.append("    { id: \"USR-2\", name: \"Admin User\",   initials: \"AU\", email: \"admin@pps.app\",")
    lines.append("      role: \"Admin\", projects: [\"All Projects\"],")
    lines.append("      status: \"Active\", lastActive: \"Today\" }")
    lines.append("  ]")
    lines.append("")
    lines.append("};")

    output = "\n".join(lines)
    if tail:
        output += tail
    else:
        output += "\n\n// =============================================================\n"
        output += "// PROJECT COLLECTION\n"
        output += "// =============================================================\n"
        output += "const PROJECTS = [PPS];\n\n"
        output += "// =============================================================\n"
        output += "// GLOBAL USERS\n"
        output += "// =============================================================\n"
        output += "const USERS = [\n"
        output += "  { id: \"USR-1\", name: \"John Tabler\", initials: \"JT\", email: \"jtabler@sdccd.edu\",\n"
        output += "    role: \"Work Dad\", projects: [\"SDCCD 1450 Frazee\"], status: \"Active\", lastActive: \"Today\" },\n"
        output += "  { id: \"USR-2\", name: \"Admin User\", initials: \"AU\", email: \"admin@pps.app\",\n"
        output += "    role: \"Admin\", projects: [\"All Projects\"], status: \"Active\", lastActive: \"Today\" }\n"
        output += "];\n"

    JS_PATH.write_text(output, encoding="utf-8")
    print(f"Written: {JS_PATH} ({JS_PATH.stat().st_size:,} bytes)")

if __name__ == "__main__":
    main()
